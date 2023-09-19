# import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime
from typing import List
import io


class ClassInfo:
    Days_Map = {
        "M": "MO",
        "T": "TU",
        "W": "WE",
        "TH": "TH",
        "F": "FR",
    }

    start_date_map = {
        "MO": 0,
        "TU": 1,
        "WE": 2,
        "TH": 3,
        "FR": 4,
    }

    def __init__(
        self,
        course_name=None,
        start_date=None,
        end_date=None,
        days=None,
        start_time=None,
        end_time=None,
        location=None,
    ) -> None:
        self.course_name = course_name
        self.start_date = start_date
        self.end_date = end_date
        self.days = days
        self.start_time = start_time
        self.end_time = end_time
        self.location = location

    def get_days(self, days):
        day = []
        for i in range(len(days)):
            if days[i : i + 2] == "TH":
                day.append("TH")
            elif days[i] != "H":
                day.append(self.Days_Map.get(days[i]))
        return sorted(day, key=lambda x: ["MO", "TU", "WE", "TH", "FR"].index(x))

    def get_start_and_end_date(self):
        start_date_map = {
            "MO": 0,
            "TU": 1,
            "WE": 2,
            "TH": 3,
            "FR": 4,
        }

        self.start_date: datetime
        self.end_date: datetime

        dayofweek = self.start_date.weekday()
        # print(dayofweek)
        if dayofweek != start_date_map[self.days[0]]:
            # print(self.days)
            Min_Day_Difference = 7
            for day in self.days:
                days_differece = (7 + (start_date_map[day] - dayofweek)) % 7
                if days_differece < Min_Day_Difference:
                    Min_Day_Difference = days_differece
                    # if Min_Day_Difference == 1: break
            import calendar
            max_day = calendar.monthrange(self.start_date.year, self.start_date.month)[1]
            if (self.start_date.day + Min_Day_Difference) > max_day:
                self.start_date = self.start_date.replace(month=self.start_date.month + 1, day=self.start_date.day + Min_Day_Difference - max_day)
            else:
                self.start_date = self.start_date.replace(day=self.start_date.day + Min_Day_Difference)
            # print(Min_Day_Difference)

        self.start_date = self.start_date.strftime("%Y%m%d")
        self.end_date = self.end_date.strftime("%Y%m%d")

    def iowrite(self, outfile: io.StringIO):
        outfile.write(
f"""BEGIN:VEVENT
DTSTART;TZID=America/Detroit:{self.start_date}T{self.start_time}
DTEND;TZID=America/Detroit:{self.start_date}T{self.end_time}
RRULE:FREQ=WEEKLY;UNTIL={self.end_date};BYDAY={",".join(self.days)}
LOCATION:{self.location}
SEQUENCE:0
STATUS:CONFIRMED
SUMMARY:{self.course_name}
END:VEVENT
""")
        



# Function to check if a string can be parsed as a date
def is_date(string):
    try:
        datetime.strptime(string, "%m/%d/%Y")
        return True
    except ValueError:
        return False

# Function to check if a string can be parsed as a time
def is_time(string):
    try:
        datetime.strptime(string, "%I:%M %p")
        return True
    except ValueError:
        return False

def get_cInfos(sheet: Worksheet):
    classes: List[ClassInfo] = []
    max_row = sheet.max_row

    for i in range(4, max_row + 1):

        if sheet.cell(row=i, column=1).value is None:
            continue

        date_time_location = sheet.cell(row=i, column=8).value
        dtl_lines = [j for j in date_time_location.split("\n") if j]
        for dtl_line in dtl_lines:
            sections = [dtl.strip() for dtl in dtl_line.split("|")]
            cInfo = ClassInfo()
            has_day = False
            for section in sections:
                section: List[str] = [s.strip() for s in section.split("-") if s]
                # section = [s.strip() for s in section.split("-") if s]
                if len(section) == 2 and is_date(section[0]):
                    continue
                if len(section) == 2 and is_time(section[0]):
                    section: List[str] = [
                        datetime.strptime(time, "%I:%M %p").strftime("%H%M%S")
                        for time in section
                    ]
                    cInfo.start_time = section[0]
                    cInfo.end_time = section[1]
                    continue
                if len(section) == 1 and has_day == False:
                    cInfo.days = cInfo.get_days(section[0])
                    has_day = True
                    continue
                if len(section) == 2:
                    cInfo.location = " - ".join(section)
                    continue
            
            cInfo.course_name = sheet.cell(row=i, column=5).value

            cInfo.start_date = sheet.cell(row=i, column=11).value
            cInfo.end_date = sheet.cell(row=i, column=12).value
            cInfo.get_start_and_end_date()

            classes.append(cInfo)
    classes.append(ClassInfo())
    return classes

def io_write(outfile: io.StringIO, classes: List[ClassInfo]):
    outfile.write(
"""BEGIN:VCALENDAR
VERSION:2.0
CALSCALE:GREGORIAN
""")
    i = 0
    while i < len(classes) - 1:
        # for i in range(len(classes)-1):
        # if course not the same, write to ics
        if classes[i].course_name != classes[i + 1].course_name:
            classes[i].iowrite(outfile)
            i += 1
            continue

        # if the course has different days, write to ics separately
        if classes[i].days != classes[i + 1].days:
            classes[i].iowrite(outfile)
            # print(f"{i+1}: {classes[i].course_name}: {classes[i].days}")
            i += 1
            continue

        # if the course has different location join the location and write to ics, then skip the next course
        if (
            classes[i].location != classes[i + 1].location
            and classes[i].start_time == classes[i + 1].start_time
            and classes[i].end_time == classes[i + 1].end_time
        ):
            classes[i].location = f"{classes[i].location}, {classes[i+1].location}"
            classes[i].iowrite(outfile)
            # print(f"{i+1}: {classes[i].course_name}: {classes[i].location}")
            # skip the next course
            i += 2
            continue

    # write the end of the ics file
    outfile.write("END:VCALENDAR")